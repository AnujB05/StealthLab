"""
Tests for the two document-processing skills built for the medical
report → Excel client use case.

`extract_pdf_text` and `build_excel` are tested against real files (a
real generated PDF, a real written-and-reread xlsx) since they're
deterministic and need no API access. The LLM-structuring half of
`pdf_extraction.py` is tested with a scripted agent -- real extraction
quality against actual medical-report language is unverified until a
real model runs it, same as everywhere else in this project.
"""
from __future__ import annotations

import asyncio
import json
import os
import tempfile

import pytest

from app.skills.excel_generation import _coerce_value, build_excel
from app.skills.pdf_extraction import extract_pdf_text, make_pdf_field_extraction_skill


def _make_test_pdf(path: str) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    doc = SimpleDocTemplate(path, pagesize=letter)
    styles = getSampleStyleSheet()
    doc.build([
        Paragraph("Patient: Test Patient", styles["Normal"]),
        Table([
            ["Test", "Result", "Unit"],
            ["Hemoglobin", "13.2", "g/dL"],
        ]),
    ])


# --- PDF text extraction (real, deterministic) ---

def test_extracts_text_and_table_from_a_real_pdf():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "report.pdf")
        _make_test_pdf(path)
        text = extract_pdf_text(path)

    assert "Test Patient" in text
    assert "Hemoglobin" in text
    assert "13.2" in text
    assert "g/dL" in text


def test_empty_pdf_raises_clearly_rather_than_returning_nothing():
    """
    A PDF with no extractable text (typically a scan) is a real, distinct
    failure from "the report legitimately had no fields" -- must not be
    silently returned as an empty success.
    """
    from reportlab.pdfgen import canvas

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "blank.pdf")
        c = canvas.Canvas(path)
        c.save()  # a real PDF with zero pages of text
        with pytest.raises(ValueError, match="no extractable text"):
            extract_pdf_text(path)


# --- LLM structuring (scripted agent, wiring under test) ---

class ScriptedExtractionAgent:
    agent_id, model_id, family = "test_agent", "mock", "mockfam"

    def __init__(self, payload):
        self.payload = payload
        self.received_prompt = None

    async def respond(self, system, user):
        self.received_prompt = user
        return json.dumps(self.payload)


def test_extraction_skill_passes_real_pdf_text_to_the_model():
    """The model must see the actual extracted text, not a placeholder."""
    agent = ScriptedExtractionAgent({"fields": {"X": {"value": 1, "unit": ""}}})
    skill = make_pdf_field_extraction_skill(agent)

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "report.pdf")
        _make_test_pdf(path)
        result = asyncio.run(skill({"pdf_path": path}))

    assert "Test Patient" in agent.received_prompt
    assert "Hemoglobin" in agent.received_prompt
    assert result["fields"] == {"X": {"value": 1, "unit": ""}}


def test_extraction_skill_surfaces_model_warnings():
    agent = ScriptedExtractionAgent({"fields": {}, "warnings": ["page 2 was unreadable"]})
    skill = make_pdf_field_extraction_skill(agent)

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "report.pdf")
        _make_test_pdf(path)
        result = asyncio.run(skill({"pdf_path": path}))

    assert result["warnings"] == ["page 2 was unreadable"]


# --- Excel generation ---

def test_coerce_value_makes_numeric_strings_real_numbers():
    """The actual correctness requirement: a number must not be a string."""
    assert _coerce_value("13.2") == 13.2
    assert isinstance(_coerce_value("13.2"), float)
    assert _coerce_value("94") == 94
    assert isinstance(_coerce_value("94"), int)
    assert _coerce_value(13.2) == 13.2  # already-numeric passes through unchanged


def test_coerce_value_leaves_real_text_as_text():
    assert _coerce_value("Jane Doe") == "Jane Doe"
    assert isinstance(_coerce_value("Jane Doe"), str)


def test_build_excel_produces_a_real_readable_file_with_correct_types():
    from openpyxl import load_workbook

    fields = {
        "Patient Name": {"value": "Jane Doe", "unit": ""},
        "Hemoglobin": {"value": "13.2", "unit": "g/dL"},  # string in, must become numeric out
    }

    with tempfile.TemporaryDirectory() as d:
        result = asyncio.run(build_excel({"fields": fields, "output_dir": d}))
        assert result["field_count"] == 2
        assert result["numeric_field_count"] == 1

        wb = load_workbook(result["file_path"])
        ws = wb.active
        rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}

        assert rows["Patient Name"][1] == "Jane Doe"
        assert isinstance(rows["Hemoglobin"][1], float)
        assert rows["Hemoglobin"][1] == 13.2


def test_build_excel_accepts_bare_value_shape_not_only_dict_shape():
    """Not tied to one specific upstream producer's exact output shape."""
    with tempfile.TemporaryDirectory() as d:
        result = asyncio.run(build_excel({"fields": {"X": 42}, "output_dir": d}))
        assert result["field_count"] == 1


def test_reference_range_never_gets_coerced_to_a_number():
    """
    A range like "13.0-17.0" or "< 200" must stay text. Coercing it would
    either fail silently (harmless) or, worse, mangle it into something
    wrong -- ranges are never a single number and must never be treated
    as one, unlike the Value column.
    """
    from openpyxl import load_workbook

    fields = {
        "Hemoglobin": {"value": 14.2, "unit": "g/dL", "reference_range": "13.0-17.0"},
        "Total Cholesterol": {"value": 184, "unit": "mg/dL", "reference_range": "< 200"},
    }
    with tempfile.TemporaryDirectory() as d:
        result = asyncio.run(build_excel({"fields": fields, "output_dir": d}))
        wb = load_workbook(result["file_path"])
        ws = wb.active
        rows = {r[0].value: r for r in ws.iter_rows(min_row=2)}

        assert rows["Hemoglobin"][3].value == "13.0-17.0"
        assert isinstance(rows["Hemoglobin"][3].value, str)
        assert rows["Total Cholesterol"][3].value == "< 200"
        # The value column itself is unaffected by this change
        assert isinstance(rows["Hemoglobin"][1].value, float)
