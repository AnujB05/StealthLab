"""
Structured fields in, a real .xlsx file out.

The one correctness detail that actually matters for this client's
stated need ("with all fields... and its numbers correctly"): a numeric
value written as a string looks identical in Excel until someone tries
to sum a column or sort it numerically, and then it silently doesn't
work. `_coerce_value` makes the actual cell type match the data's real
type, not just its string representation.
"""
from __future__ import annotations

import os
import uuid
from typing import Any


def _coerce_value(value: Any):
    """
    Write real numbers as numbers. A value that's already a JSON number
    passes through; a numeric-looking string ("142.5", "-3") is
    converted rather than left as text, since an upstream model call
    (see pdf_extraction.py) could plausibly emit either depending on
    exactly how it formatted its JSON.
    """
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        try:
            if "." in stripped:
                return float(stripped)
            return int(stripped)
        except ValueError:
            return value
    return value


async def build_combined_excel(input_data: dict[str, Any]) -> dict[str, Any]:
    """
    Multiple reports' fields in, one wide comparison table out:
    Field | Value (report A) | Value (report B) | ... | Unit | Reference Range

    `input_data["reports"]` is `{display_name: fields_dict}` for however
    many reports were successfully extracted. Field sets don't have to
    match across reports -- the union of all field names becomes the row
    list, and a report missing a given field just leaves that cell blank
    rather than raising, since two different lab reports legitimately
    won't always test the same panel.

    Unit and reference range are taken from whichever report has them
    first for that field name. For the same field name (e.g.
    "Hemoglobin") these should agree across reports of the same kind, so
    one shared pair of columns is the right shape, not a repeated pair
    per report.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    reports: dict[str, dict[str, Any]] = input_data["reports"]
    output_dir = input_data.get("output_dir", "/tmp")
    os.makedirs(output_dir, exist_ok=True)

    report_names = list(reports.keys())

    # Union of field names, first-seen order across reports -- not
    # alphabetical, so the row order still roughly follows the source
    # report's own structure rather than scrambling it.
    field_names: list[str] = []
    seen: set[str] = set()
    for fields in reports.values():
        for name in fields:
            if name not in seen:
                seen.add(name)
                field_names.append(name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Fields"

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    headers = (
        ["Field"] + [f"Value ({name})" for name in report_names] + ["Unit", "Reference Range"]
    )
    for col, title in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=title).font = header_font

    unit_col = len(report_names) + 2
    ref_col = len(report_names) + 3

    row = 2
    for field_name in field_names:
        ws.cell(row=row, column=1, value=field_name).font = body_font

        unit, reference_range = "", ""
        for offset, report_name in enumerate(report_names):
            detail = reports[report_name].get(field_name)
            value = None
            if detail is not None:
                if isinstance(detail, dict):
                    value = _coerce_value(detail.get("value"))
                    if not unit:
                        unit = detail.get("unit", "") or ""
                    if not reference_range:
                        reference_range = str(detail.get("reference_range", "") or "")
                else:
                    value = _coerce_value(detail)
            ws.cell(row=row, column=2 + offset, value=value).font = body_font

        ws.cell(row=row, column=unit_col, value=unit).font = body_font
        ws.cell(row=row, column=ref_col, value=reference_range).font = body_font
        row += 1

    col_letters = [chr(ord("A") + i) for i in range(len(headers))]
    widths = [36] + [18] * len(report_names) + [10, 16]
    for letter, width in zip(col_letters, widths):
        ws.column_dimensions[letter].width = width

    filename = f"combined-{uuid.uuid4().hex[:8]}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)

    return {
        "file_path": path,
        "report_count": len(report_names),
        "field_count": len(field_names),
    }


async def build_excel(input_data: dict[str, Any]) -> dict[str, Any]:
    """
    `input_data["fields"]` is the shape `pdf_extraction.py`'s skill
    produces: `{"field name": {"value": ..., "unit": "..."}}`. Also
    accepts a bare `{"field name": value}` shape, so this skill isn't
    tied to one specific upstream producer.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    fields: dict[str, Any] = input_data["fields"]
    output_dir = input_data.get("output_dir", "/tmp")
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Fields"

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    for col, title in enumerate(("Field", "Value", "Unit", "Reference Range"), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    row = 2
    numeric_count = 0
    for name, detail in fields.items():
        if isinstance(detail, dict):
            raw_value = detail.get("value")
            unit = detail.get("unit", "") or ""
            # Deliberately never coerced: "13.0-17.0" or "< 200" would
            # either fail to parse as a number (harmless) or, worse,
            # silently mangle into something wrong if a range ever looked
            # numeric-ish on its own. Reference ranges stay text, always.
            reference_range = str(detail.get("reference_range", "") or "")
        else:
            raw_value, unit, reference_range = detail, "", ""

        value = _coerce_value(raw_value)
        if isinstance(value, (int, float)):
            numeric_count += 1

        ws.cell(row=row, column=1, value=name).font = body_font
        ws.cell(row=row, column=2, value=value).font = body_font
        ws.cell(row=row, column=3, value=unit).font = body_font
        ws.cell(row=row, column=4, value=reference_range).font = body_font
        row += 1

    for col, width in zip("ABCD", (36, 16, 10, 16)):
        ws.column_dimensions[col].width = width

    filename = f"extracted-{uuid.uuid4().hex[:8]}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)

    return {
        "file_path": path,
        "field_count": len(fields),
        "numeric_field_count": numeric_count,
    }
